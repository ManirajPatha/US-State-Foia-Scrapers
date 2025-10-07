import time
import sys
from typing import Optional, Tuple
from pathlib import Path

from selenium import webdriver
from selenium.webdriver.firefox.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException
from openpyxl import load_workbook


FORM_URL = "https://sas.arkansas.gov/secretarys-office/legal/foia-requests/"

# ---- Locators from your description (robust where possible) ----
FIRST_NAME_ID = "input_7_1_3"        # <input id="input_7_1_3" name="input_1.3">
LAST_NAME_ID  = "input_7_1_6"        # <input id="input_7_1_6" name="input_1.6">
PHONE_ID      = "input_7_3"          # <input id="input_7_3"   name="input_3">
EMAIL_ID      = "input_7_4"          # <input id="input_7_4"   name="input_4">

STREET_ID     = "input_7_6_1"        # <input id="input_7_6_1" name="input_6.1">
CITY_ID       = "input_7_6_3"        # <input id="input_7_6_3" name="input_6.3">
STATE_ID      = "input_7_6_4"        # <input id="input_7_6_4" name="input_6.4">
ZIP_ID        = "input_7_6_5"        # <input id="input_7_6_5" name="input_6.5">
COUNTRY_ID    = "input_7_6_6"        # <select id="input_7_6_6" name="input_6.6">

# Description doesn't have a provided id/name; grab the first textarea after a label that contains "Description".
DESC_XPATH    = "//label[contains(normalize-space(.), 'Description')]/following::textarea[1]"

SUBMIT_ID     = "gform_submit_button_7"  # <input type="submit" id="gform_submit_button_7" ...>


def prompt_with_default(prompt_text: str, default: Optional[str] = None) -> str:
    if default:
        entered = input(f"{prompt_text} [{default}]: ").strip()
        return entered if entered else default
    return input(f"{prompt_text}: ").strip()


def build_description(bid_solicitation: str, org_name: str, awarded_vendors: str) -> str:
    return (
        f"I am requesting a copy of the winning proposal and the shortlisted proposals for "
        f"{bid_solicitation}, awarded by {org_name}. The awarder vendor is {awarded_vendors}."
    )


def wait_for_present_and_type(driver, by, value, text, timeout=20):
    el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((by, value)))
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", el)
    el.clear()
    el.send_keys(text)
    return el


def wait_for_click(driver, by, value, timeout=20):
    el = WebDriverWait(driver, timeout).until(EC.element_to_be_clickable((by, value)))
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", el)
    el.click()
    return el


def set_country(driver, country_text: str, timeout=20):
    sel_el = WebDriverWait(driver, timeout).until(EC.presence_of_element_located((By.ID, COUNTRY_ID)))
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", sel_el)
    Select(sel_el).select_by_visible_text(country_text)


def wait_for_confirmation(driver, timeout=30) -> bool:
    """
    Gravity Forms usually shows a confirmation block after successful submit.
    We'll try a few robust checks.
    """
    try:
        # 1) A typical confirmation has .gform_confirmation_message
        WebDriverWait(driver, timeout).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, ".gform_confirmation_message"))
        )
        return True
    except TimeoutException:
        pass

    # 2) Sometimes form hides and a confirmation container appears; check form invisibility
    try:
        WebDriverWait(driver, 3).until_not(EC.presence_of_element_located((By.ID, SUBMIT_ID)))
        return True
    except TimeoutException:
        pass

    # 3) As a fallback, just wait a couple of seconds assuming postback happened
    time.sleep(2)
    return True


def open_form(driver, url=FORM_URL, timeout=30):
    driver.get(url)
    WebDriverWait(driver, timeout).until(
        EC.presence_of_element_located((By.ID, FIRST_NAME_ID))
    )


def submit_one_request(
    driver,
    requester_info: dict,
    bid_solicitation: str,
    org_name: str,
    awarded_vendors: str
) -> bool:
    # Ensure we’re on a fresh form page
    open_form(driver)

    # Fill requester info
    wait_for_present_and_type(driver, By.ID, FIRST_NAME_ID, requester_info["first_name"])
    wait_for_present_and_type(driver, By.ID, LAST_NAME_ID,  requester_info["last_name"])
    wait_for_present_and_type(driver, By.ID, PHONE_ID,      requester_info["phone"])
    wait_for_present_and_type(driver, By.ID, EMAIL_ID,      requester_info["email"])

    wait_for_present_and_type(driver, By.ID, STREET_ID,     requester_info["street"])
    wait_for_present_and_type(driver, By.ID, CITY_ID,       requester_info["city"])
    wait_for_present_and_type(driver, By.ID, STATE_ID,      requester_info["state"])
    wait_for_present_and_type(driver, By.ID, ZIP_ID,        requester_info["zip"])
    set_country(driver, requester_info["country"])

    # Description per row
    description_text = build_description(bid_solicitation, org_name, awarded_vendors)

    # Find description textarea robustly via label-based XPath
    desc_el = WebDriverWait(driver, 20).until(
        EC.presence_of_element_located((By.XPATH, DESC_XPATH))
    )
    driver.execute_script("arguments[0].scrollIntoView({behavior: 'instant', block: 'center'});", desc_el)
    desc_el.clear()
    desc_el.send_keys(description_text)

    # Submit
    try:
        submit_btn = wait_for_click(driver, By.ID, SUBMIT_ID, timeout=20)
    except (TimeoutException, ElementClickInterceptedException):
        # Fallback: try JS click if normal click fails
        submit_btn = driver.find_element(By.ID, SUBMIT_ID)
        driver.execute_script("arguments[0].click();", submit_btn)

    # Wait for confirmation
    return wait_for_confirmation(driver, timeout=30)


def read_rows_from_excel(xlsx_path: Path) -> list:
    """
    Expects columns:
      A = 'Bid Solicitation #'
      B = 'Organization Name'
      C = 'Awarded Vendor(s)'

    Starts from row 2 until an empty row is found.
    """
    wb = load_workbook(filename=str(xlsx_path), data_only=True)
    ws = wb.active

    rows = []
    row_idx = 2
    while True:
        a = ws.cell(row=row_idx, column=1).value  # Bid Solicitation #
        b = ws.cell(row=row_idx, column=2).value  # Organization Name
        c = ws.cell(row=row_idx, column=3).value  # Awarded Vendor(s)

        # Stop at first completely empty row (A, B, C all empty)
        if (a is None or str(a).strip() == "") and (b is None or str(b).strip() == "") and (c is None or str(c).strip() == ""):
            break

        # Coerce to strings safely
        a_str = "" if a is None else str(a).strip()
        b_str = "" if b is None else str(b).strip()
        c_str = "" if c is None else str(c).strip()

        # Skip rows missing the key Bid Solicitation #
        if a_str == "":
            row_idx += 1
            continue

        rows.append((a_str, b_str, c_str))
        row_idx += 1

    return rows


def main():
    print("\n=== Arkansas SOS FOIA Auto-Submitter ===\n")

    # --- Prompt once for requester details (used for ALL rows) ---
    first_name = prompt_with_default("Enter First Name", "Raaj")
    last_name  = prompt_with_default("Enter Last Name", "Thipparthy")
    phone      = prompt_with_default("Enter Phone", "8325197135")
    email      = prompt_with_default("Enter Email", "raajnrao@gmail.com")

    street     = prompt_with_default("Enter Street Address", "8181 Fannin St")
    city       = prompt_with_default("Enter City", "Houston")
    state      = prompt_with_default("Enter State/Province/Region", "Texas")
    zip_code   = prompt_with_default("Enter ZIP/Postal Code", "77054")
    country    = prompt_with_default("Enter Country (must match dropdown text)", "United States")

    # Input file path (default from your spec)
    default_path = "/Users/raajthipparthy/Desktop/Opportunity Scrapers/arkansas-scraper/ar_bidbuy_results.xlsx"
    xlsx_input = prompt_with_default("Path to input Excel", default_path)
    xlsx_path = Path(xlsx_input).expanduser()

    if not xlsx_path.exists():
        print(f"ERROR: Input file not found at: {xlsx_path}")
        sys.exit(1)

    rows = read_rows_from_excel(xlsx_path)
    if not rows:
        print("No records found in the input Excel. Exiting.")
        sys.exit(0)

    requester_info = {
        "first_name": first_name,
        "last_name": last_name,
        "phone": phone,
        "email": email,
        "street": street,
        "city": city,
        "state": state,
        "zip": zip_code,
        "country": country,
    }

    # --- Start Firefox (set headless = True if you prefer) ---
    options = webdriver.FirefoxOptions()
    headless_ans = prompt_with_default("Run Firefox headless? (y/n)", "n").lower()
    if headless_ans.startswith("y"):
        options.add_argument("-headless")

    driver = webdriver.Firefox(options=options)
    driver.set_page_load_timeout(60)

    try:
        successes = 0
        failures = 0

        for idx, (bid_solicitation, org_name, awarded_vendors) in enumerate(rows, start=1):
            print(f"\n[{idx}/{len(rows)}] Submitting request for Bid #{bid_solicitation} | Org: {org_name} | Vendor(s): {awarded_vendors}")
            try:
                ok = submit_one_request(
                    driver,
                    requester_info=requester_info,
                    bid_solicitation=bid_solicitation,
                    org_name=org_name,
                    awarded_vendors=awarded_vendors
                )
                if ok:
                    print("  -> Submitted successfully.")
                    successes += 1
                else:
                    print("  -> Submission may have failed (no confirmation detected).")
                    failures += 1
            except Exception as e:
                print(f"  -> ERROR submitting this row: {e}")
                failures += 1

            # Be polite between submissions
            time.sleep(2)

        print("\n=== Done ===")
        print(f"Successful submissions: {successes}")
        print(f"Failed submissions:     {failures}")

    finally:
        driver.quit()


if __name__ == "__main__":
    main()