from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
import time
import os
import logging
import zipfile
from openpyxl import Workbook, load_workbook

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')

chrome_options = webdriver.ChromeOptions()
download_dir = os.path.join(os.getcwd(), "downloads")
if not os.path.exists(download_dir):
    os.makedirs(download_dir)
    os.chmod(download_dir, 0o755)

chrome_options.add_experimental_option("prefs", {
    "download.default_directory": download_dir,
    "download.prompt_for_download": False,
    "download.directory_upgrade": True,
    "safebrowsing.enabled": True
})

try:
    driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=chrome_options)
    logging.info("WebDriver initialized successfully")
except Exception as e:
    logging.error(f"Failed to initialize WebDriver: {e}")
    raise

excel_file = os.path.join(download_dir, "scraped_data.xlsx")
if os.path.exists(excel_file):
    wb = load_workbook(excel_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.append(["Solicitation Number", "Department", "Status", "Opening Date", "Posted Date", "Owner", "Description"])

def wait_for_download(directory, timeout=30):
    """Wait for download to complete"""
    seconds = 0
    while seconds < timeout:
        time.sleep(1)
        if not any([filename.endswith(".crdownload") or filename.endswith(".tmp") for filename in os.listdir(directory)]):
            return True
        seconds += 1
    return False

def scrape_opportunity(driver, opportunity_link, download_dir, excel_file):
    """Scrape a single opportunity and download attachments"""
    
    opportunity_link.click()
    time.sleep(3)
    
    try:
        # Scrape data
        solicitation_number = driver.find_element(By.ID, "evp_solicitationnbr").get_attribute("value")
        department = driver.find_element(By.ID, "owningbusinessunit_name").get_attribute("value")
        status = driver.find_element(By.ID, "statuscode").text
        opening_date = driver.find_element(By.ID, "evp_opendate_datepicker_description").get_attribute("value")
        posted_date = driver.find_element(By.ID, "evp_posteddate_datepicker_description").get_attribute("value")
        owner = driver.find_element(By.ID, "ownerid_name").get_attribute("value")
        description = driver.find_element(By.ID, "evp_description").get_attribute("value")
        
        logging.info(f"Scraped data for: {solicitation_number}")
        
        attachment_folder = os.path.join(download_dir, f"{solicitation_number}_attachments")
        if not os.path.exists(attachment_folder):
            os.makedirs(attachment_folder)
        
        attachment_links = driver.find_elements(By.CSS_SELECTOR, "div.attachment a[href*='/_entity/annotation']")
        
        for link in attachment_links:
            try:
                filename = link.text.split(" (")[0]
                logging.info(f"Downloading: {filename}")
                
                link.click()
                wait_for_download(download_dir)
                
                # Move the downloaded file to the attachment folder
                downloaded_files = [f for f in os.listdir(download_dir) if os.path.isfile(os.path.join(download_dir, f)) and not f.endswith('.zip') and not f.endswith('.xlsx')]
                if downloaded_files:
                    latest_file = max([os.path.join(download_dir, f) for f in downloaded_files], key=os.path.getctime)
                    new_path = os.path.join(attachment_folder, os.path.basename(latest_file))
                    os.rename(latest_file, new_path)
                    logging.info(f"Moved {latest_file} to {new_path}")
                
                time.sleep(1)
            except Exception as e:
                logging.error(f"Error downloading attachment: {e}")
        
        zip_filename = os.path.join(download_dir, f"{solicitation_number}_attachments.zip")
        with zipfile.ZipFile(zip_filename, 'w') as zipf:
            for root, dirs, files in os.walk(attachment_folder):
                for file in files:
                    file_path = os.path.join(root, file)
                    zipf.write(file_path, os.path.basename(file_path))
        
        logging.info(f"Created zip file: {zip_filename}")
        
        ws.append([solicitation_number, department, status, opening_date, posted_date, owner, description])
        wb.save(excel_file)
        
    except Exception as e:
        logging.error(f"Error scraping opportunity: {e}")

try:
    logging.info("Navigating to website")
    driver.get("https://evp.nc.gov/solicitations/")

    logging.info("Clicking filter header")
    filter_header = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "filter-header"))
    )
    filter_header.click()

    logging.info("Selecting 'Last Year' in Posted Date")
    posted_date_dropdown = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.ID, "dropdownfilter_2"))
    )
    posted_date_dropdown.click()
    last_year_option = driver.find_element(By.XPATH, "//select[@id='dropdownfilter_2']/option[@value='5']")
    last_year_option.click()

    logging.info("Unchecking 'Open' in Solicitation Status")
    open_checkbox = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@name='3' and @value='0']"))
    )
    if open_checkbox.is_selected():
        open_checkbox.click()

    logging.info("Checking 'Awarded' in Solicitation Status")
    awarded_checkbox = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@name='3' and @value='2']"))
    )
    if not awarded_checkbox.is_selected():
        awarded_checkbox.click()

    logging.info("Clicking Apply button")
    apply_button = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.CSS_SELECTOR, "button.btn-entitylist-filter-submit"))
    )
    apply_button.click()

    logging.info("Waiting for filtered results to load...")
    time.sleep(5)
    
    # Wait for table to update with filtered data
    WebDriverWait(driver, 15).until(
        EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped tbody tr"))
    )
    time.sleep(3)
    
    # Verify we have "Awarded" status entries
    try:
        first_row_status = driver.find_element(By.CSS_SELECTOR, "table.table-striped tbody tr:first-child td[data-attribute='statuscode']").text
        logging.info(f"First row status after filter: {first_row_status}")
        
        if "Awarded" not in first_row_status:
            logging.warning("Filter may not have applied correctly. First row status is not 'Awarded'")
            time.sleep(5)
    except Exception as e:
        logging.warning(f"Could not verify filter status: {e}")

    # Main pagination loop
    current_page = 1
    
    while True:
        logging.info(f"Processing page {current_page}")
        
        WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped tbody tr"))
        )
        
        opportunity_rows = driver.find_elements(By.CSS_SELECTOR, "table.table-striped tbody tr")
        num_opportunities = len(opportunity_rows)
        
        logging.info(f"Found {num_opportunities} opportunities on page {current_page}")
        
        # Process each opportunity on the current page
        for i in range(num_opportunities):
            if i > 0:
                logging.info("Navigating back using browser back button")
                driver.back()
                time.sleep(3)
                
                WebDriverWait(driver, 10).until(
                    EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped tbody tr"))
                )
            
            opportunity_rows = driver.find_elements(By.CSS_SELECTOR, "table.table-striped tbody tr")
            opportunity_link = opportunity_rows[i].find_element(By.CSS_SELECTOR, "a.details-link")
            logging.info(f"Processing opportunity {i+1}/{num_opportunities}: {opportunity_link.text}")
            
            scrape_opportunity(driver, opportunity_link, download_dir, excel_file)
            logging.info(f"Completed opportunity {i+1}/{num_opportunities}")
        
        logging.info(f"Completed page {current_page}")
        
        # Check if there's a next page
        try:
            logging.info("Navigating back to list for pagination check")
            driver.back()
            time.sleep(3)
            
            WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.CSS_SELECTOR, "table.table-striped tbody tr"))
            )
            
            next_button = driver.find_element(By.CSS_SELECTOR, "a.entity-pager-next-link")
            next_button_class = next_button.get_attribute("class") or ""
            next_button_style = next_button.get_attribute("style") or ""
            
            if "disabled" in next_button_class or "display: none" in next_button_style:
                logging.info("No more pages available. Scraping complete.")
                break
            
            logging.info(f"Moving to page {current_page + 1}")
            next_button.click()
            time.sleep(3)
            current_page += 1
            
        except Exception as e:
            logging.info(f"No more pages or error finding next button: {e}")
            break

    logging.info("All opportunities scraped successfully")

except Exception as e:
    logging.error(f"An error occurred: {e}")
finally:
    wb.save(excel_file)
    logging.info(f"Excel file saved: {excel_file}")
    logging.info("Closing browser")
    driver.quit()